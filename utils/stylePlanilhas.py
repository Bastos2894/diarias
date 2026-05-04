from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# faz a formatação da planilha de saída, deixando ela mais apresentável e fácil de ler
def style_planilhas(arquivo_saida):

    wb = load_workbook(arquivo_saida)
    ws = wb.active

    # estilos
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    thick_border = Border(
        bottom=Side(style="thick"),
        top=Side(style="thick"),
        left=Side(style="thick"),
        right=Side(style="thick")
    )

    # cabeçalho
    for cell in ws[1]:
        cell.font = bold_font
        cell.border = thick_border
        cell.fill = header_fill

    # congelar
    ws.freeze_panes = "A2"

    # auto width
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    # tabela
    last_row = ws.max_row
    last_col = ws.max_column
    table_range = f"A1:{get_column_letter(last_col)}{last_row}"

    table = Table(displayName="TabelaDados", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    ws.auto_filter.ref = table_range

    wb.save(arquivo_saida)